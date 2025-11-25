"""
LEADERSHIP EXCEL REPORT - HOSPITAL SIMILARITY BY TOWER
=======================================================

Creates an Excel workbook with 6 tabs (one per embedding tower):
1. Procedure/Treatment
2. Diagnosis
3. Demographics
4. Place of Service
5. Cost Category
6. Totality (Overall) - Table 1 only

Sampling: 2 hospitals per label (including unlabeled), selected from
hospitals with many procedure codes. Same sample used across all tabs.

TOP 5 ALTERNATIVES: Based on OVERALL similarity (full 1046-dim embedding),
same 5 alternatives shown across all tabs.

For each primary hospital:
- Table 1: Top 5 alternatives with overlap stats
- Table 2: Code-level drilldown (except Totality)

Author: AI Assistant
Date: 2025
"""

import pandas as pd
import numpy as np
import pickle
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =============================================================================
# LOAD DATA
# =============================================================================

print("\n" + "="*80)
print("LOADING DATA")
print("="*80)

with open('pin_to_label.pkl', 'rb') as f:
    pin_to_label = pickle.load(f)
print(f"Labels: {len(pin_to_label)}")

embeddings_df = pd.read_parquet('final_me2vec_all_towers_1046d.parquet')
print(f"Embeddings: {embeddings_df.shape}")

procedure_df = pd.read_parquet('procedure_df.parquet')
print(f"Procedure data: {procedure_df.shape}")

diagnosis_df = pd.read_parquet('diagnosis_df.parquet')
print(f"Diagnosis data: {diagnosis_df.shape}")

demo_df = pd.read_parquet('demo_df.parquet')
print(f"Demographics: {demo_df.shape}")

place_df = pd.read_parquet('place_df.parquet')
print(f"Place: {place_df.shape}")

cost_df = pd.read_parquet('cost_df.parquet')
print(f"Cost: {cost_df.shape}")

pin_df = pd.read_parquet('pin_df.parquet')
print(f"PIN summary: {pin_df.shape}")

pin_names_df = pd.read_parquet('all_pin_names.parquet')
all_pins_with_embeddings = embeddings_df['PIN'].values
pin_names_df = pin_names_df[pin_names_df['PIN'].isin(all_pins_with_embeddings)]
print(f"PIN names: {pin_names_df.shape}")

pin_to_name = dict(zip(pin_names_df['PIN'], pin_names_df['PIN_name']))

prov_spl_df = pd.read_parquet('prov_spl.parquet')
print(f"Provider specialty: {prov_spl_df.shape}")
pin_to_specialty = dict(zip(prov_spl_df['PIN'], prov_spl_df['srv_spclty_ctg_cd']))

# Load code descriptions
code_desc_df = pd.read_parquet('code_desc_grouped.parquet')
print(f"Code descriptions: {code_desc_df.shape}")

# Aggregate multiple descriptions per code
code_to_desc = (
    code_desc_df
    .groupby('code')['code_desc']
    .apply(lambda x: ', '.join(x.unique()))
    .to_dict()
)
print(f"Unique codes with descriptions: {len(code_to_desc)}")

cnp_matrix = np.load('cnp_matrix.npy')
print(f"CNP matrix shape: {cnp_matrix.shape}")

# =============================================================================
# PREPARE DATA STRUCTURES
# =============================================================================

print("\n" + "="*80)
print("PREPARING DATA STRUCTURES")
print("="*80)

all_pins_list = embeddings_df['PIN'].values.tolist()
pin_to_idx = {pin: idx for idx, pin in enumerate(all_pins_list)}

emb_cols = [col for col in embeddings_df.columns if col != 'PIN']
pin_to_emb = dict(zip(embeddings_df['PIN'], embeddings_df[emb_cols].values.tolist()))

# Tower dimensions
tower_dims = {
    'procedure': (0, 512),
    'diagnosis': (512, 1024),
    'demographics': (1024, 1029),
    'place': (1029, 1033),
    'cost': (1033, 1044),
}

# Process procedure data
procedure_summary = procedure_df.groupby('PIN').agg({
    'code': lambda x: set(x),
    'claims': lambda x: dict(zip(procedure_df.loc[x.index, 'code'], x))
}).reset_index()
procedure_summary.columns = ['PIN', 'codes', 'code_to_claims']
pin_to_procedure_codes = dict(zip(procedure_summary['PIN'], procedure_summary['codes']))
pin_to_procedure_claims = dict(zip(procedure_summary['PIN'], procedure_summary['code_to_claims']))

# Process diagnosis data
diagnosis_summary = diagnosis_df.groupby('PIN').agg({
    'code': lambda x: set(x),
    'claims': lambda x: dict(zip(diagnosis_df.loc[x.index, 'code'], x))
}).reset_index()
diagnosis_summary.columns = ['PIN', 'codes', 'code_to_claims']
pin_to_diagnosis_codes = dict(zip(diagnosis_summary['PIN'], diagnosis_summary['codes']))
pin_to_diagnosis_claims = dict(zip(diagnosis_summary['PIN'], diagnosis_summary['code_to_claims']))

# Process demographics, place, cost as indexed dataframes
demo_df_indexed = demo_df.set_index('PIN')
demo_cols = list(demo_df_indexed.columns)

place_df_indexed = place_df.set_index('PIN')
place_cols = list(place_df_indexed.columns)

cost_df_indexed = cost_df.set_index('PIN')
cost_cols = list(cost_df_indexed.columns)

print(f"Demographics columns: {demo_cols}")
print(f"Place columns: {place_cols}")
print(f"Cost columns: {cost_cols}")

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def cosine_similarity_manual(vec_a, vec_b):
    vec_a = np.array(vec_a)
    vec_b = np.array(vec_b)
    dot = np.dot(vec_a, vec_b)
    norm_a = np.linalg.norm(vec_a)
    norm_b = np.linalg.norm(vec_b)
    return dot / (norm_a * norm_b + 1e-8)

def compute_tower_similarity(emb_a, emb_b, tower_name):
    start_idx, end_idx = tower_dims[tower_name]
    tower_a = np.array(emb_a)[start_idx:end_idx]
    tower_b = np.array(emb_b)[start_idx:end_idx]
    return cosine_similarity_manual(tower_a, tower_b)

def get_top_5_alternatives(primary_pin):
    """Get top 5 alternatives based on OVERALL similarity (full 1046-dim embedding)."""
    if primary_pin not in pin_to_idx:
        return []
    
    primary_idx = pin_to_idx[primary_pin]
    primary_emb = pin_to_emb.get(primary_pin)
    
    if primary_emb is None:
        return []
    
    similarities = []
    for alt_pin in all_pins_list:
        if alt_pin == primary_pin:
            continue
        alt_emb = pin_to_emb.get(alt_pin)
        if alt_emb is None:
            continue
        
        # OVERALL similarity using full embedding
        overall_sim = cosine_similarity_manual(primary_emb, alt_emb)
        
        # CNP scores for reference
        alt_idx = pin_to_idx[alt_pin]
        cnp_score = cnp_matrix[primary_idx, alt_idx]
        
        similarities.append((alt_pin, overall_sim, cnp_score))
    
    # Sort by OVERALL similarity
    similarities.sort(key=lambda x: x[1], reverse=True)
    return similarities[:5]

def compute_overlap_stats(primary_pin, alt_pin, pin_to_codes, pin_to_claims):
    """Compute overlap statistics between primary and alternative."""
    primary_codes = pin_to_codes.get(primary_pin, set())
    alt_codes = pin_to_codes.get(alt_pin, set())
    primary_claims = pin_to_claims.get(primary_pin, {})
    alt_claims = pin_to_claims.get(alt_pin, {})
    
    overlapping_codes = primary_codes & alt_codes
    non_overlapping_primary = primary_codes - alt_codes
    
    overlap_count = len(overlapping_codes)
    
    primary_claims_overlap = sum(primary_claims.get(c, 0) for c in overlapping_codes)
    primary_claims_non_overlap = sum(primary_claims.get(c, 0) for c in non_overlapping_primary)
    
    alt_total_claims = sum(alt_claims.values())
    alt_claims_non_overlap = sum(alt_claims.get(c, 0) for c in (alt_codes - primary_codes))
    alt_pct_non_overlap = (alt_claims_non_overlap / alt_total_claims * 100) if alt_total_claims > 0 else 0
    
    return {
        'overlap_count': overlap_count,
        'primary_claims_overlap': primary_claims_overlap,
        'primary_claims_non_overlap': primary_claims_non_overlap,
        'alt_pct_non_overlap': alt_pct_non_overlap
    }

def compute_column_overlap_stats(primary_pin, alt_pin, df_indexed, cols):
    """Compute overlap stats for column-based data (demographics, place, cost)."""
    if primary_pin not in df_indexed.index or alt_pin not in df_indexed.index:
        return {'overlap_count': 0, 'primary_claims_overlap': 0, 
                'primary_claims_non_overlap': 0, 'alt_pct_non_overlap': 0}
    
    primary_vals = df_indexed.loc[primary_pin]
    alt_vals = df_indexed.loc[alt_pin]
    
    primary_codes = set(c for c in cols if primary_vals[c] > 0)
    alt_codes = set(c for c in cols if alt_vals[c] > 0)
    
    overlapping = primary_codes & alt_codes
    non_overlapping_primary = primary_codes - alt_codes
    
    primary_claims_overlap = sum(primary_vals[c] for c in overlapping)
    primary_claims_non_overlap = sum(primary_vals[c] for c in non_overlapping_primary)
    
    alt_total = sum(alt_vals)
    alt_non_overlap = sum(alt_vals[c] for c in (alt_codes - primary_codes))
    alt_pct_non_overlap = (alt_non_overlap / alt_total * 100) if alt_total > 0 else 0
    
    return {
        'overlap_count': len(overlapping),
        'primary_claims_overlap': primary_claims_overlap,
        'primary_claims_non_overlap': primary_claims_non_overlap,
        'alt_pct_non_overlap': alt_pct_non_overlap
    }

# =============================================================================
# SAMPLE HOSPITALS
# =============================================================================

print("\n" + "="*80)
print("SAMPLING HOSPITALS")
print("="*80)

procedure_code_counts = {pin: len(codes) for pin, codes in pin_to_procedure_codes.items()}

label_to_pins = {}
for pin, label in pin_to_label.items():
    if pin in pin_to_idx:
        if label not in label_to_pins:
            label_to_pins[label] = []
        label_to_pins[label].append(pin)

unlabeled_pins = [pin for pin in all_pins_list if pin not in pin_to_label]
if unlabeled_pins:
    label_to_pins['unlabeled'] = unlabeled_pins

print(f"\nLabels found: {list(label_to_pins.keys())}")
for label, pins in label_to_pins.items():
    print(f"  {label}: {len(pins)} hospitals")

sampled_hospitals = []
random.seed(42)

for label, pins in label_to_pins.items():
    pins_with_procedures = [p for p in pins if p in procedure_code_counts]
    pins_with_procedures.sort(key=lambda p: procedure_code_counts.get(p, 0), reverse=True)
    
    top_percentile = max(1, len(pins_with_procedures) // 5)
    top_pins = pins_with_procedures[:top_percentile]
    
    if len(top_pins) >= 2:
        selected = random.sample(top_pins, 2)
    elif len(top_pins) == 1:
        selected = top_pins
    elif len(pins_with_procedures) >= 2:
        selected = random.sample(pins_with_procedures, 2)
    elif len(pins_with_procedures) == 1:
        selected = pins_with_procedures
    else:
        selected = []
    
    for pin in selected:
        sampled_hospitals.append({
            'pin': pin,
            'label': label,
            'name': pin_to_name.get(pin, 'Unknown'),
            'specialty': pin_to_specialty.get(pin, 'Unknown'),
            'procedure_code_count': procedure_code_counts.get(pin, 0)
        })

print(f"\nSampled {len(sampled_hospitals)} hospitals:")
for h in sampled_hospitals:
    print(f"  {h['label']}: {h['pin']} - {h['name'][:50]} ({h['procedure_code_count']} procedure codes)")

# =============================================================================
# PRECOMPUTE TOP 5 ALTERNATIVES (OVERALL SIMILARITY)
# =============================================================================

print("\n" + "="*80)
print("COMPUTING TOP 5 ALTERNATIVES (OVERALL SIMILARITY)")
print("="*80)

hospital_to_top5 = {}
for hospital in sampled_hospitals:
    primary_pin = hospital['pin']
    top_5 = get_top_5_alternatives(primary_pin)
    hospital_to_top5[primary_pin] = top_5
    print(f"  {primary_pin}: Found {len(top_5)} alternatives")

# =============================================================================
# EXCEL STYLING
# =============================================================================

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row_num, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def style_data_cell(ws, row_num, col_num):
    cell = ws.cell(row=row_num, column=col_num)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="right")

def auto_adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# =============================================================================
# CREATE EXCEL WORKBOOK
# =============================================================================

print("\n" + "="*80)
print("CREATING EXCEL WORKBOOK")
print("="*80)

wb = Workbook()
wb.remove(wb.active)

# Tower configurations
tower_configs = [
    {
        'name': 'Procedure_Treatment',
        'tower': 'procedure',
        'pin_to_codes': pin_to_procedure_codes,
        'pin_to_claims': pin_to_procedure_claims,
        'use_code_desc': True,
        'df_indexed': None,
        'cols': None
    },
    {
        'name': 'Diagnosis',
        'tower': 'diagnosis',
        'pin_to_codes': pin_to_diagnosis_codes,
        'pin_to_claims': pin_to_diagnosis_claims,
        'use_code_desc': True,
        'df_indexed': None,
        'cols': None
    },
    {
        'name': 'Demographics',
        'tower': 'demographics',
        'pin_to_codes': None,
        'pin_to_claims': None,
        'use_code_desc': False,
        'df_indexed': demo_df_indexed,
        'cols': demo_cols
    },
    {
        'name': 'Place_of_Service',
        'tower': 'place',
        'pin_to_codes': None,
        'pin_to_claims': None,
        'use_code_desc': False,
        'df_indexed': place_df_indexed,
        'cols': place_cols
    },
    {
        'name': 'Cost_Category',
        'tower': 'cost',
        'pin_to_codes': None,
        'pin_to_claims': None,
        'use_code_desc': False,
        'df_indexed': cost_df_indexed,
        'cols': cost_cols
    },
    {
        'name': 'Totality',
        'tower': None,  # No specific tower
        'pin_to_codes': None,
        'pin_to_claims': None,
        'use_code_desc': False,
        'df_indexed': None,
        'cols': None,
        'table1_only': True  # Only Table 1 for Totality
    }
]

for config in tower_configs:
    print(f"\nCreating tab: {config['name']}")
    ws = wb.create_sheet(title=config['name'])
    
    current_row = 1
    is_table1_only = config.get('table1_only', False)
    
    for hospital in sampled_hospitals:
        primary_pin = hospital['pin']
        primary_name = hospital['name']
        primary_label = hospital['label']
        
        print(f"  Processing: {primary_pin} ({primary_label})")
        
        # Get PRECOMPUTED top 5 alternatives (same across all tabs)
        top_5 = hospital_to_top5.get(primary_pin, [])
        
        if not top_5:
            print(f"    No alternatives found, skipping")
            continue
        
        # =================================================================
        # PRIMARY HOSPITAL HEADER
        # =================================================================
        ws.cell(row=current_row, column=1, value=f"Primary Hospital: {primary_name}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"PIN: {primary_pin} | Label: {primary_label} | Specialty: {hospital['specialty']}")
        current_row += 2
        
        # =================================================================
        # TABLE 1: ALTERNATIVES SUMMARY
        # =================================================================
        ws.cell(row=current_row, column=1, value="Table 1: Top 5 Alternatives Summary")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
        
        # Table 1 headers
        table1_headers = [
            'Rank', 'Alternative PIN', 'Alternative Name', 'Specialty',
            'Overall Similarity', 'CNP Score', 'Overlapping Codes',
            'Primary Claims (Overlap)', 'Primary Claims (Non-Overlap)',
            '% Alt Claims Non-Overlap'
        ]
        
        for col, header in enumerate(table1_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        style_header_row(ws, current_row, 1, len(table1_headers))
        current_row += 1
        
        # Table 1 data
        alt_pins_for_table2 = []
        for rank, (alt_pin, overall_sim, cnp_score) in enumerate(top_5, 1):
            alt_pins_for_table2.append(alt_pin)
            
            # Get overlap stats based on tower type
            if config['pin_to_codes'] is not None:
                overlap_stats = compute_overlap_stats(
                    primary_pin, alt_pin, 
                    config['pin_to_codes'], config['pin_to_claims']
                )
            elif config['df_indexed'] is not None:
                overlap_stats = compute_column_overlap_stats(
                    primary_pin, alt_pin,
                    config['df_indexed'], config['cols']
                )
            else:
                # Totality - use procedure as proxy
                overlap_stats = compute_overlap_stats(
                    primary_pin, alt_pin, 
                    pin_to_procedure_codes, pin_to_procedure_claims
                )
            
            row_data = [
                rank,
                alt_pin,
                pin_to_name.get(alt_pin, 'Unknown')[:50],
                pin_to_specialty.get(alt_pin, 'Unknown'),
                round(overall_sim, 4),
                round(cnp_score, 6),
                overlap_stats['overlap_count'],
                overlap_stats['primary_claims_overlap'],
                overlap_stats['primary_claims_non_overlap'],
                round(overlap_stats['alt_pct_non_overlap'], 2)
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value)
                style_data_cell(ws, current_row, col)
            current_row += 1
        
        current_row += 2
        
        # =================================================================
        # TABLE 2: CODE-LEVEL DRILLDOWN (Skip for Totality)
        # =================================================================
        if is_table1_only:
            current_row += 1
            continue
        
        ws.cell(row=current_row, column=1, value="Table 2: Code-Level Drilldown")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
        
        # Build Table 2 headers
        table2_headers = ['Code', 'Description', 'Primary Claims', 'Primary %']
        for i in range(1, 6):
            table2_headers.append(f'Alt{i} Claims')
            table2_headers.append(f'Alt{i} %')
        
        for col, header in enumerate(table2_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        style_header_row(ws, current_row, 1, len(table2_headers))
        current_row += 1
        
        # Table 2 data
        if config['pin_to_codes'] is not None:
            # Code-based data (procedure, diagnosis)
            primary_codes = config['pin_to_codes'].get(primary_pin, set())
            primary_claims = config['pin_to_claims'].get(primary_pin, {})
            primary_total = sum(primary_claims.values())
            
            alt_claims_list = []
            alt_totals = []
            for alt_pin in alt_pins_for_table2:
                claims = config['pin_to_claims'].get(alt_pin, {})
                alt_claims_list.append(claims)
                alt_totals.append(sum(claims.values()))
            
            # Collect all codes
            all_codes = set(primary_codes)
            for alt_pin in alt_pins_for_table2:
                all_codes.update(config['pin_to_codes'].get(alt_pin, set()))
            
            # Sort by primary claims descending, limit to 50
            all_codes = sorted(all_codes, key=lambda c: primary_claims.get(c, 0), reverse=True)[:50]
            
            for code in all_codes:
                desc = code_to_desc.get(code, code) if config['use_code_desc'] else code
                
                primary_claim_count = primary_claims.get(code, 0)
                primary_pct = (primary_claim_count / primary_total * 100) if primary_total > 0 else 0
                
                row_data = [code, str(desc)[:100], primary_claim_count, round(primary_pct, 2)]
                
                for i, alt_pin in enumerate(alt_pins_for_table2):
                    alt_claim_count = alt_claims_list[i].get(code, 0)
                    alt_pct = (alt_claim_count / alt_totals[i] * 100) if alt_totals[i] > 0 else 0
                    row_data.append(alt_claim_count)
                    row_data.append(round(alt_pct, 2))
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                    style_data_cell(ws, current_row, col)
                current_row += 1
        
        else:
            # Column-based data (demographics, place, cost)
            cols = config['cols']
            df_indexed = config['df_indexed']
            
            if primary_pin in df_indexed.index:
                primary_vals = df_indexed.loc[primary_pin]
                primary_total = sum(primary_vals)
            else:
                primary_vals = pd.Series(0, index=cols)
                primary_total = 0
            
            alt_vals_list = []
            alt_totals = []
            for alt_pin in alt_pins_for_table2:
                if alt_pin in df_indexed.index:
                    vals = df_indexed.loc[alt_pin]
                else:
                    vals = pd.Series(0, index=cols)
                alt_vals_list.append(vals)
                alt_totals.append(sum(vals))
            
            # Sort columns by primary value descending
            sorted_cols = sorted(cols, key=lambda c: primary_vals.get(c, 0), reverse=True)
            
            for col_name in sorted_cols:
                desc = col_name  # For demographics/place/cost, description = code
                
                primary_val = primary_vals.get(col_name, 0)
                primary_pct = (primary_val / primary_total * 100) if primary_total > 0 else 0
                
                row_data = [col_name, desc, primary_val, round(primary_pct, 2)]
                
                for i, alt_pin in enumerate(alt_pins_for_table2):
                    alt_val = alt_vals_list[i].get(col_name, 0)
                    alt_pct = (alt_val / alt_totals[i] * 100) if alt_totals[i] > 0 else 0
                    row_data.append(alt_val)
                    row_data.append(round(alt_pct, 2))
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                    style_data_cell(ws, current_row, col)
                current_row += 1
        
        current_row += 3  # Space before next hospital
    
    auto_adjust_column_width(ws)

# =============================================================================
# SAVE WORKBOOK
# =============================================================================

output_file = 'leadership_hospital_similarity_report.xlsx'
wb.save(output_file)
print(f"\n" + "="*80)
print(f"EXCEL REPORT SAVED: {output_file}")
print("="*80)
print(f"\nSummary:")
print(f"  - {len(sampled_hospitals)} primary hospitals sampled")
print(f"  - 6 tabs created (one per tower)")
print(f"  - Top 5 alternatives based on OVERALL similarity (same across tabs)")
print(f"  - Totality tab: Table 1 only (no code drilldown)")
