"""
LEADERSHIP EXCEL REPORT - PROCEDURE TAB WITH PIVOTTABLES
=========================================================

Creates an Excel workbook with:
1. Raw data tabs (source for pivots)
2. Procedure_Treatment tab with PivotTables + Slicer

Slicer allows selecting 1 of 32 sampled providers, and both tables refresh.

NOTE: On first open, you must click "Data" > "Refresh All" (Ctrl+Alt+F5)
to populate the PivotTables.

Author: AI Assistant
Date: 2025
"""

import pandas as pd
import numpy as np
import pickle
import random
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.pivot.table import PivotTable, PivotTableStyleInfo
from openpyxl.pivot.cache import CacheDefinition, CacheField, CacheSource
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# LOAD DATA
# =============================================================================

print("\n" + "="*80)
print("LOADING DATA")
print("="*80)

with open('pin_to_label.pkl', 'rb') as f:
    pin_to_label = pickle.load(f)
print(f"Labels: {len(pin_to_label)}")

embeddings_df = pd.read_parquet('final_me2vec_all_towers_1046d.parquet')
print(f"Embeddings: {embeddings_df.shape}")

procedure_df = pd.read_parquet('procedure_df.parquet')
print(f"Procedure data: {procedure_df.shape}")

pin_names_df = pd.read_parquet('all_pin_names.parquet')
all_pins_with_embeddings = embeddings_df['PIN'].values
pin_names_df = pin_names_df[pin_names_df['PIN'].isin(all_pins_with_embeddings)]
print(f"PIN names: {pin_names_df.shape}")

pin_to_name = dict(zip(pin_names_df['PIN'], pin_names_df['PIN_name']))

prov_spl_df = pd.read_parquet('prov_spl.parquet')
print(f"Provider specialty: {prov_spl_df.shape}")
pin_to_specialty = dict(zip(prov_spl_df['PIN'], prov_spl_df['srv_spclty_ctg_cd']))

# Load code descriptions
code_desc_df = pd.read_parquet('code_desc_grouped.parquet')
print(f"Code descriptions: {code_desc_df.shape}")

code_to_desc = (
    code_desc_df
    .groupby('code')['code_desc']
    .apply(lambda x: ', '.join(x.unique()))
    .to_dict()
)
print(f"Unique codes with descriptions: {len(code_to_desc)}")

cnp_matrix = np.load('cnp_matrix.npy')
print(f"CNP matrix shape: {cnp_matrix.shape}")

# =============================================================================
# PREPARE DATA STRUCTURES
# =============================================================================

print("\n" + "="*80)
print("PREPARING DATA STRUCTURES")
print("="*80)

all_pins_list = embeddings_df['PIN'].values.tolist()
pin_to_idx = {pin: idx for idx, pin in enumerate(all_pins_list)}

emb_cols = [col for col in embeddings_df.columns if col != 'PIN']
pin_to_emb = dict(zip(embeddings_df['PIN'], embeddings_df[emb_cols].values.tolist()))

# Process procedure data
procedure_summary = procedure_df.groupby('PIN').agg({
    'code': lambda x: set(x),
    'claims': lambda x: dict(zip(procedure_df.loc[x.index, 'code'], x))
}).reset_index()
procedure_summary.columns = ['PIN', 'codes', 'code_to_claims']
pin_to_procedure_codes = dict(zip(procedure_summary['PIN'], procedure_summary['codes']))
pin_to_procedure_claims = dict(zip(procedure_summary['PIN'], procedure_summary['code_to_claims']))

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def cosine_similarity_manual(vec_a, vec_b):
    vec_a = np.array(vec_a)
    vec_b = np.array(vec_b)
    dot = np.dot(vec_a, vec_b)
    norm_a = np.linalg.norm(vec_a)
    norm_b = np.linalg.norm(vec_b)
    return dot / (norm_a * norm_b + 1e-8)

def get_top_5_alternatives(primary_pin):
    """Get top 5 alternatives based on OVERALL similarity."""
    if primary_pin not in pin_to_idx:
        return []
    
    primary_idx = pin_to_idx[primary_pin]
    primary_emb = pin_to_emb.get(primary_pin)
    
    if primary_emb is None:
        return []
    
    similarities = []
    for alt_pin in all_pins_list:
        if alt_pin == primary_pin:
            continue
        alt_emb = pin_to_emb.get(alt_pin)
        if alt_emb is None:
            continue
        
        overall_sim = cosine_similarity_manual(primary_emb, alt_emb)
        alt_idx = pin_to_idx[alt_pin]
        cnp_score = cnp_matrix[primary_idx, alt_idx]
        
        similarities.append((alt_pin, overall_sim, cnp_score))
    
    similarities.sort(key=lambda x: x[1], reverse=True)
    return similarities[:5]

def compute_overlap_stats(primary_pin, alt_pin):
    """Compute overlap statistics for procedures."""
    primary_codes = pin_to_procedure_codes.get(primary_pin, set())
    alt_codes = pin_to_procedure_codes.get(alt_pin, set())
    primary_claims = pin_to_procedure_claims.get(primary_pin, {})
    alt_claims = pin_to_procedure_claims.get(alt_pin, {})
    
    overlapping_codes = primary_codes & alt_codes
    non_overlapping_primary = primary_codes - alt_codes
    
    overlap_count = len(overlapping_codes)
    primary_claims_overlap = sum(primary_claims.get(c, 0) for c in overlapping_codes)
    primary_claims_non_overlap = sum(primary_claims.get(c, 0) for c in non_overlapping_primary)
    
    alt_total_claims = sum(alt_claims.values())
    alt_claims_non_overlap = sum(alt_claims.get(c, 0) for c in (alt_codes - primary_codes))
    alt_pct_non_overlap = (alt_claims_non_overlap / alt_total_claims * 100) if alt_total_claims > 0 else 0
    
    return {
        'overlap_count': overlap_count,
        'primary_claims_overlap': primary_claims_overlap,
        'primary_claims_non_overlap': primary_claims_non_overlap,
        'alt_pct_non_overlap': alt_pct_non_overlap
    }

# =============================================================================
# SAMPLE HOSPITALS
# =============================================================================

print("\n" + "="*80)
print("SAMPLING HOSPITALS")
print("="*80)

procedure_code_counts = {pin: len(codes) for pin, codes in pin_to_procedure_codes.items()}

label_to_pins = {}
for pin, label in pin_to_label.items():
    if pin in pin_to_idx:
        if label not in label_to_pins:
            label_to_pins[label] = []
        label_to_pins[label].append(pin)

unlabeled_pins = [pin for pin in all_pins_list if pin not in pin_to_label]
if unlabeled_pins:
    label_to_pins['unlabeled'] = unlabeled_pins

print(f"\nLabels found: {list(label_to_pins.keys())}")

sampled_hospitals = []
random.seed(42)

for label, pins in label_to_pins.items():
    pins_with_procedures = [p for p in pins if p in procedure_code_counts]
    pins_with_procedures.sort(key=lambda p: procedure_code_counts.get(p, 0), reverse=True)
    
    top_percentile = max(1, len(pins_with_procedures) // 5)
    top_pins = pins_with_procedures[:top_percentile]
    
    if len(top_pins) >= 2:
        selected = random.sample(top_pins, 2)
    elif len(top_pins) == 1:
        selected = top_pins
    elif len(pins_with_procedures) >= 2:
        selected = random.sample(pins_with_procedures, 2)
    elif len(pins_with_procedures) == 1:
        selected = pins_with_procedures
    else:
        selected = []
    
    for pin in selected:
        sampled_hospitals.append({
            'pin': pin,
            'label': label,
            'name': pin_to_name.get(pin, 'Unknown'),
            'specialty': pin_to_specialty.get(pin, 'Unknown'),
            'procedure_code_count': procedure_code_counts.get(pin, 0)
        })

print(f"\nSampled {len(sampled_hospitals)} hospitals")

# =============================================================================
# PRECOMPUTE TOP 5 ALTERNATIVES
# =============================================================================

print("\n" + "="*80)
print("COMPUTING TOP 5 ALTERNATIVES")
print("="*80)

hospital_to_top5 = {}
for hospital in sampled_hospitals:
    primary_pin = hospital['pin']
    top_5 = get_top_5_alternatives(primary_pin)
    hospital_to_top5[primary_pin] = top_5
    print(f"  {primary_pin}: Found {len(top_5)} alternatives")

# =============================================================================
# BUILD RAW DATA FOR TABLE 1 (Alternatives Summary)
# =============================================================================

print("\n" + "="*80)
print("BUILDING RAW DATA - TABLE 1")
print("="*80)

table1_rows = []

for hospital in sampled_hospitals:
    primary_pin = hospital['pin']
    primary_name = hospital['name']
    primary_label = hospital['label']
    primary_specialty = hospital['specialty']
    
    # Create a display name for the slicer (more readable)
    primary_display = f"{primary_name[:40]} ({primary_pin})"
    
    top_5 = hospital_to_top5.get(primary_pin, [])
    
    for rank, (alt_pin, overall_sim, cnp_score) in enumerate(top_5, 1):
        overlap_stats = compute_overlap_stats(primary_pin, alt_pin)
        
        table1_rows.append({
            'Primary_Hospital': primary_display,
            'Primary_PIN': primary_pin,
            'Primary_Name': primary_name,
            'Primary_Label': primary_label,
            'Primary_Specialty': primary_specialty,
            'Rank': rank,
            'Alternative_PIN': alt_pin,
            'Alternative_Name': pin_to_name.get(alt_pin, 'Unknown'),
            'Alternative_Specialty': pin_to_specialty.get(alt_pin, 'Unknown'),
            'Overall_Similarity': round(overall_sim, 4),
            'CNP_Score': round(cnp_score, 6),
            'Overlapping_Codes': overlap_stats['overlap_count'],
            'Primary_Claims_Overlap': overlap_stats['primary_claims_overlap'],
            'Primary_Claims_NonOverlap': overlap_stats['primary_claims_non_overlap'],
            'Alt_Pct_NonOverlap': round(overlap_stats['alt_pct_non_overlap'], 2)
        })

table1_df = pd.DataFrame(table1_rows)
print(f"Table 1 raw data: {table1_df.shape}")

# =============================================================================
# BUILD RAW DATA FOR TABLE 2 (Code-Level Drilldown)
# =============================================================================

print("\n" + "="*80)
print("BUILDING RAW DATA - TABLE 2")
print("="*80)

table2_rows = []

for hospital in sampled_hospitals:
    primary_pin = hospital['pin']
    primary_name = hospital['name']
    primary_display = f"{primary_name[:40]} ({primary_pin})"
    
    top_5 = hospital_to_top5.get(primary_pin, [])
    alt_pins = [alt_pin for alt_pin, _, _ in top_5]
    
    # Get all codes from primary and alternatives
    primary_codes = pin_to_procedure_codes.get(primary_pin, set())
    primary_claims = pin_to_procedure_claims.get(primary_pin, {})
    primary_total = sum(primary_claims.values()) if primary_claims else 1
    
    all_codes = set(primary_codes)
    for alt_pin in alt_pins:
        all_codes.update(pin_to_procedure_codes.get(alt_pin, set()))
    
    # Sort by primary claims, limit to top 50
    all_codes = sorted(all_codes, key=lambda c: primary_claims.get(c, 0), reverse=True)[:50]
    
    for code in all_codes:
        desc = code_to_desc.get(code, code)
        if len(str(desc)) > 100:
            desc = str(desc)[:100]
        
        primary_claim_count = primary_claims.get(code, 0)
        primary_pct = round((primary_claim_count / primary_total * 100), 2) if primary_total > 0 else 0
        
        row = {
            'Primary_Hospital': primary_display,
            'Primary_PIN': primary_pin,
            'Code': code,
            'Description': desc,
            'Primary_Claims': primary_claim_count,
            'Primary_Pct': primary_pct
        }
        
        # Add alternatives
        for i, alt_pin in enumerate(alt_pins, 1):
            alt_claims = pin_to_procedure_claims.get(alt_pin, {})
            alt_total = sum(alt_claims.values()) if alt_claims else 1
            alt_claim_count = alt_claims.get(code, 0)
            alt_pct = round((alt_claim_count / alt_total * 100), 2) if alt_total > 0 else 0
            
            row[f'Alt{i}_Claims'] = alt_claim_count
            row[f'Alt{i}_Pct'] = alt_pct
        
        # Fill missing alternatives with 0
        for i in range(len(alt_pins) + 1, 6):
            row[f'Alt{i}_Claims'] = 0
            row[f'Alt{i}_Pct'] = 0
        
        table2_rows.append(row)

table2_df = pd.DataFrame(table2_rows)
print(f"Table 2 raw data: {table2_df.shape}")

# =============================================================================
# CREATE EXCEL WORKBOOK
# =============================================================================

print("\n" + "="*80)
print("CREATING EXCEL WORKBOOK")
print("="*80)

wb = Workbook()
wb.remove(wb.active)

# -----------------------------------------------------------------------------
# TAB 1: Proc_Table1_Data (Raw data for Table 1)
# -----------------------------------------------------------------------------
print("Creating Proc_Table1_Data tab...")
ws1 = wb.create_sheet(title='Proc_Table1_Data')

# Write headers
for col_idx, col_name in enumerate(table1_df.columns, 1):
    ws1.cell(row=1, column=col_idx, value=col_name)
    ws1.cell(row=1, column=col_idx).font = Font(bold=True)

# Write data
for row_idx, row in enumerate(table1_df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws1.cell(row=row_idx, column=col_idx, value=value)

# Create Excel Table
table1_range = f"A1:{get_column_letter(len(table1_df.columns))}{len(table1_df) + 1}"
table1 = Table(displayName="ProcTable1Data", ref=table1_range)
table1.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
ws1.add_table(table1)

# Adjust column widths
for col_idx in range(1, len(table1_df.columns) + 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 18

# -----------------------------------------------------------------------------
# TAB 2: Proc_Table2_Data (Raw data for Table 2)
# -----------------------------------------------------------------------------
print("Creating Proc_Table2_Data tab...")
ws2 = wb.create_sheet(title='Proc_Table2_Data')

# Write headers
for col_idx, col_name in enumerate(table2_df.columns, 1):
    ws2.cell(row=1, column=col_idx, value=col_name)
    ws2.cell(row=1, column=col_idx).font = Font(bold=True)

# Write data
for row_idx, row in enumerate(table2_df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Create Excel Table
table2_range = f"A1:{get_column_letter(len(table2_df.columns))}{len(table2_df) + 1}"
table2 = Table(displayName="ProcTable2Data", ref=table2_range)
table2.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
ws2.add_table(table2)

# Adjust column widths
for col_idx in range(1, len(table2_df.columns) + 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 15

# -----------------------------------------------------------------------------
# TAB 3: Procedure_Treatment (Instructions + View Tab)
# -----------------------------------------------------------------------------
print("Creating Procedure_Treatment tab...")
ws_main = wb.create_sheet(title='Procedure_Treatment', index=0)

# Add instructions
ws_main['A1'] = "PROCEDURE/TREATMENT ANALYSIS"
ws_main['A1'].font = Font(bold=True, size=16)

ws_main['A3'] = "HOW TO USE THIS WORKBOOK:"
ws_main['A3'].font = Font(bold=True, size=12)

ws_main['A5'] = "1. On first open: Click 'Data' > 'Refresh All' (or Ctrl+Alt+F5)"
ws_main['A6'] = "2. Go to 'Insert' > 'PivotTable' to create PivotTables from the data tabs"
ws_main['A7'] = "3. Add Slicers: Click on PivotTable > 'PivotTable Analyze' > 'Insert Slicer' > Select 'Primary_Hospital'"
ws_main['A8'] = "4. Connect Slicers to both PivotTables: Right-click Slicer > 'Report Connections'"

ws_main['A10'] = "DATA TABS:"
ws_main['A10'].font = Font(bold=True, size=12)

ws_main['A12'] = "• Proc_Table1_Data: Top 5 alternatives summary (160 rows = 32 hospitals × 5 alternatives)"
ws_main['A13'] = "• Proc_Table2_Data: Code-level drilldown (up to 50 codes per hospital)"

ws_main['A15'] = "RECOMMENDED PIVOTTABLE 1 LAYOUT (Alternatives Summary):"
ws_main['A15'].font = Font(bold=True, size=12)

ws_main['A17'] = "• Filters: Primary_Hospital (use Slicer)"
ws_main['A18'] = "• Rows: Rank"
ws_main['A19'] = "• Values: Alternative_PIN, Alternative_Name, Overall_Similarity, CNP_Score, Overlapping_Codes, Primary_Claims_Overlap"

ws_main['A21'] = "RECOMMENDED PIVOTTABLE 2 LAYOUT (Code Drilldown):"
ws_main['A21'].font = Font(bold=True, size=12)

ws_main['A23'] = "• Filters: Primary_Hospital (connect to same Slicer)"
ws_main['A24'] = "• Rows: Code, Description"
ws_main['A25'] = "• Values: Primary_Claims, Primary_Pct, Alt1_Claims, Alt1_Pct, Alt2_Claims, Alt2_Pct, etc."

ws_main['A27'] = "SAMPLED HOSPITALS:"
ws_main['A27'].font = Font(bold=True, size=12)

# List sampled hospitals
row = 29
for i, hospital in enumerate(sampled_hospitals, 1):
    ws_main[f'A{row}'] = f"{i}. {hospital['name'][:50]} (PIN: {hospital['pin']}, Label: {hospital['label']})"
    row += 1

# Adjust column width
ws_main.column_dimensions['A'].width = 100

# =============================================================================
# SAVE WORKBOOK
# =============================================================================

output_file = 'procedure_pivot_report.xlsx'
wb.save(output_file)

print(f"\n" + "="*80)
print(f"EXCEL REPORT SAVED: {output_file}")
print("="*80)
print(f"\nContents:")
print(f"  1. Procedure_Treatment - Instructions tab")
print(f"  2. Proc_Table1_Data - Raw data for Table 1 ({len(table1_df)} rows)")
print(f"  3. Proc_Table2_Data - Raw data for Table 2 ({len(table2_df)} rows)")
print(f"\nNEXT STEPS:")
print(f"  1. Open the file in Excel")
print(f"  2. Click 'Data' > 'Refresh All'")
print(f"  3. Create PivotTables from the data tabs")
print(f"  4. Add Slicer for 'Primary_Hospital' to filter both tables")
